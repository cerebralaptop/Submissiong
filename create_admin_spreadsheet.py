#!/usr/bin/env python3
"""
Green Star Admin Spreadsheet Generator
Creates a professional Excel file for managing Green Star credit data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill_dark = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")  # Green Star green
header_fill_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
header_fill_purple = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
header_fill_amber = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
header_fill_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

wrap_alignment = Alignment(wrap_text=True, vertical='top')
center_alignment = Alignment(horizontal='center', vertical='center')

# Category colors for conditional formatting
category_colors = {
    "Responsible": "FEF3C7",  # amber-100
    "Healthy": "DCFCE7",      # green-100
    "Resilient": "DBEAFE",    # blue-100
    "Positive": "FEE2E2",     # red-100
    "Places": "F3E8FF",       # purple-100
    "People": "FCE7F3",       # pink-100
    "Nature": "D1FAE5",       # emerald-100
    "Leadership": "E0E7FF"    # indigo-100
}

# ============================================================
# SHEET 1: Instructions
# ============================================================
ws_instructions = wb.active
ws_instructions.title = "Instructions"
ws_instructions.sheet_properties.tabColor = "2D5A27"

instructions = [
    ["GREEN STAR BUILDINGS v1.1 - ADMIN SPREADSHEET"],
    [""],
    ["HOW TO USE THIS SPREADSHEET"],
    [""],
    ["This spreadsheet manages all content for the Green Star Submission Assistant app."],
    ["Edit the data in each sheet, then import it back into the app."],
    [""],
    ["SHEETS OVERVIEW:"],
    [""],
    ["1. Categories    - The 8 main category groups (Responsible, Healthy, etc.)"],
    ["2. Credits       - All credits with their basic info and category assignment"],
    ["3. Levels        - Performance levels (Minimum/Credit/Exceptional) for each credit"],
    ["4. Criteria      - Detailed criteria/requirements for each level"],
    ["5. Documentation - Required documents for each credit"],
    ["6. Templates     - Project templates with pre-selected credits"],
    ["7. Synergies     - Credit relationships and synergies"],
    [""],
    ["TIPS:"],
    ["- Use the dropdown menus where available (Category, Level, Yes/No fields)"],
    ["- Don't change Credit IDs - they link data across sheets"],
    ["- Add new rows at the bottom of each sheet"],
    ["- Keep formatting consistent"],
    ["- Export to the app after making changes"],
    [""],
    ["COLOR CODING:"],
    ["- Green headers = Main data sheets"],
    ["- Blue headers = Reference/lookup sheets"],
    ["- Yellow cells = Dropdown selections available"],
    [""],
    ["CONTACT: [Your admin contact info here]"],
]

for row_idx, row in enumerate(instructions, 1):
    cell = ws_instructions.cell(row=row_idx, column=1, value=row[0] if row else "")
    if row_idx == 1:
        cell.font = Font(bold=True, size=18, color="2D5A27")
    elif row_idx in [3, 8, 19, 26]:
        cell.font = Font(bold=True, size=12)

ws_instructions.column_dimensions['A'].width = 80

# ============================================================
# SHEET 2: Categories
# ============================================================
ws_categories = wb.create_sheet("Categories")
ws_categories.sheet_properties.tabColor = "1E40AF"

cat_headers = ["Category ID", "Category Name", "Description", "Icon (FontAwesome)", "Color Theme", "Display Order"]
for col, header in enumerate(cat_headers, 1):
    cell = ws_categories.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_blue
    cell.border = thin_border
    cell.alignment = center_alignment

categories_data = [
    ["responsible", "Responsible", "Responsible practices in construction and development", "fa-shield-halved", "amber", 1],
    ["healthy", "Healthy", "Creating healthy indoor environments", "fa-heart-pulse", "green", 2],
    ["resilient", "Resilient", "Building resilience to climate change", "fa-shield", "blue", 3],
    ["positive", "Positive", "Positive environmental outcomes", "fa-bolt", "red", 4],
    ["places", "Places", "Creating great places for communities", "fa-map-location-dot", "purple", 5],
    ["people", "People", "Supporting people and communities", "fa-users", "pink", 6],
    ["nature", "Nature", "Protecting and enhancing nature", "fa-leaf", "emerald", 7],
    ["leadership", "Leadership", "Demonstrating industry leadership", "fa-trophy", "indigo", 8],
]

for row_idx, row_data in enumerate(categories_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_categories.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx == 2:
            cell.fill = PatternFill(start_color=category_colors.get(value, "FFFFFF"), end_color=category_colors.get(value, "FFFFFF"), fill_type="solid")

# Set column widths
ws_categories.column_dimensions['A'].width = 15
ws_categories.column_dimensions['B'].width = 18
ws_categories.column_dimensions['C'].width = 50
ws_categories.column_dimensions['D'].width = 22
ws_categories.column_dimensions['E'].width = 15
ws_categories.column_dimensions['F'].width = 15

# ============================================================
# SHEET 3: Credits
# ============================================================
ws_credits = wb.create_sheet("Credits")
ws_credits.sheet_properties.tabColor = "2D5A27"

credit_headers = ["Credit ID", "Credit Name", "Category", "Short Description", "Full Description", "Is Required?", "Max Points", "Display Order"]
for col, header in enumerate(credit_headers, 1):
    cell = ws_credits.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_dark
    cell.border = thin_border
    cell.alignment = center_alignment

# Add data validation for Category column
category_list = '"Responsible,Healthy,Resilient,Positive,Places,People,Nature,Leadership"'
cat_validation = DataValidation(type="list", formula1=category_list, allow_blank=False)
cat_validation.error = "Please select a valid category"
cat_validation.errorTitle = "Invalid Category"
cat_validation.prompt = "Select a category"
cat_validation.promptTitle = "Category"
ws_credits.add_data_validation(cat_validation)
cat_validation.add('C2:C200')

# Add data validation for Is Required column
yes_no_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
yes_no_validation.error = "Please select Yes or No"
ws_credits.add_data_validation(yes_no_validation)
yes_no_validation.add('F2:F200')

# Credits data
credits_data = [
    ["resp-01", "Responsible Construction", "Responsible", "Sustainable construction practices", "Demonstrates responsible construction practices including waste management, pollution prevention, and sustainable material sourcing.", "Yes", 3, 1],
    ["resp-02", "Metering and Monitoring", "Responsible", "Building performance monitoring", "Implementation of comprehensive metering and monitoring systems to track building performance.", "No", 2, 2],
    ["resp-03", "Adaptation and Resilience", "Responsible", "Climate adaptation strategies", "Strategies to adapt to changing climate conditions and improve building resilience.", "No", 3, 3],
    ["resp-04", "Building Information", "Responsible", "Documentation and handover", "Comprehensive building information and documentation for occupants and facility managers.", "No", 2, 4],
    ["resp-05", "Commitment to Performance", "Responsible", "Performance verification", "Commitment to verify and maintain building performance over time.", "No", 3, 5],

    ["health-01", "Clean Air", "Healthy", "Indoor air quality", "Ensuring high indoor air quality through ventilation and material selection.", "Yes", 3, 1],
    ["health-02", "Lighting Comfort", "Healthy", "Visual comfort and daylight", "Providing comfortable lighting conditions with good daylight access.", "No", 2, 2],
    ["health-03", "Thermal Comfort", "Healthy", "Temperature and humidity control", "Maintaining comfortable thermal conditions for occupants.", "No", 2, 3],
    ["health-04", "Acoustic Comfort", "Healthy", "Noise control", "Controlling noise levels for occupant comfort and productivity.", "No", 2, 4],
    ["health-05", "Access to Nature", "Healthy", "Biophilic design", "Providing visual and physical access to nature.", "No", 2, 5],

    ["resil-01", "Climate Resilience", "Resilient", "Climate risk assessment", "Assessment and mitigation of climate-related risks.", "Yes", 3, 1],
    ["resil-02", "Flood Resilience", "Resilient", "Flood risk management", "Managing flood risks through design and planning.", "No", 2, 2],
    ["resil-03", "Heat Resilience", "Resilient", "Urban heat mitigation", "Strategies to reduce urban heat island effect.", "No", 2, 3],

    ["pos-01", "Energy Source", "Positive", "Renewable energy use", "Use of renewable energy sources to power the building.", "Yes", 4, 1],
    ["pos-02", "Energy Use", "Positive", "Energy efficiency", "Reducing energy consumption through efficient design.", "Yes", 4, 2],
    ["pos-03", "Upfront Carbon Reduction", "Positive", "Embodied carbon", "Reducing embodied carbon in construction materials.", "Yes", 4, 3],
    ["pos-04", "Operational Waste", "Positive", "Waste management", "Managing operational waste through reduction and recycling.", "No", 2, 4],
    ["pos-05", "Water Use", "Positive", "Water efficiency", "Reducing water consumption through efficient fixtures and systems.", "No", 3, 5],

    ["place-01", "Movement and Place", "Places", "Sustainable transport", "Supporting sustainable transport options and creating quality places.", "Yes", 3, 1],
    ["place-02", "Community Benefits", "Places", "Community engagement", "Providing benefits to the local community.", "No", 2, 2],
    ["place-03", "Culture and Heritage", "Places", "Heritage preservation", "Respecting and integrating local culture and heritage.", "No", 2, 3],

    ["people-01", "Inclusive Construction Practices", "People", "Social procurement", "Inclusive hiring and procurement practices during construction.", "Yes", 3, 1],
    ["people-02", "Health and Wellbeing", "People", "Occupant wellbeing", "Promoting health and wellbeing of building occupants.", "No", 2, 2],
    ["people-03", "Accessibility", "People", "Universal design", "Ensuring accessibility for people of all abilities.", "No", 2, 3],

    ["nature-01", "Impacts to Nature", "Nature", "Environmental protection", "Minimizing negative impacts on local ecology.", "Yes", 3, 1],
    ["nature-02", "Biodiversity Enhancement", "Nature", "Ecological improvement", "Enhancing biodiversity on and around the site.", "No", 3, 2],
    ["nature-03", "Sustainable Sites", "Nature", "Site selection", "Selecting and developing sites sustainably.", "No", 2, 3],

    ["lead-01", "Market Transformation", "Leadership", "Industry leadership", "Demonstrating leadership and innovation in the industry.", "Yes", 4, 1],
    ["lead-02", "Communication and Engagement", "Leadership", "Stakeholder engagement", "Engaging stakeholders and communicating sustainability goals.", "No", 2, 2],
    ["lead-03", "Innovation", "Leadership", "Innovative solutions", "Implementing innovative sustainability solutions.", "No", 3, 3],
]

for row_idx, row_data in enumerate(credits_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_credits.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx in [4, 5] else Alignment(vertical='top')
        # Color code by category
        if col_idx == 3 and value in category_colors:
            cell.fill = PatternFill(start_color=category_colors[value], end_color=category_colors[value], fill_type="solid")
        # Highlight required credits
        if col_idx == 6 and value == "Yes":
            cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # yellow highlight

# Set column widths
ws_credits.column_dimensions['A'].width = 12
ws_credits.column_dimensions['B'].width = 30
ws_credits.column_dimensions['C'].width = 15
ws_credits.column_dimensions['D'].width = 30
ws_credits.column_dimensions['E'].width = 60
ws_credits.column_dimensions['F'].width = 12
ws_credits.column_dimensions['G'].width = 12
ws_credits.column_dimensions['H'].width = 14

# Freeze top row
ws_credits.freeze_panes = 'A2'

# ============================================================
# SHEET 4: Performance Levels
# ============================================================
ws_levels = wb.create_sheet("Levels")
ws_levels.sheet_properties.tabColor = "2D5A27"

level_headers = ["Credit ID", "Level Type", "Points", "Summary", "Description"]
for col, header in enumerate(level_headers, 1):
    cell = ws_levels.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_dark
    cell.border = thin_border
    cell.alignment = center_alignment

# Add data validation for Level Type
level_validation = DataValidation(type="list", formula1='"Minimum Expectation,Credit Achievement,Exceptional Performance"', allow_blank=False)
level_validation.error = "Please select a valid level type"
ws_levels.add_data_validation(level_validation)
level_validation.add('B2:B500')

# Sample levels data (abbreviated - would include all credits)
levels_data = [
    ["resp-01", "Minimum Expectation", 0, "Basic compliance", "Meet basic responsible construction requirements"],
    ["resp-01", "Credit Achievement", 2, "Good practices", "Implement good responsible construction practices"],
    ["resp-01", "Exceptional Performance", 3, "Best in class", "Demonstrate industry-leading responsible construction"],

    ["health-01", "Minimum Expectation", 0, "Basic air quality", "Meet minimum ventilation and air quality standards"],
    ["health-01", "Credit Achievement", 2, "Enhanced air quality", "Exceed minimum air quality requirements"],
    ["health-01", "Exceptional Performance", 3, "Premium air quality", "Achieve premium indoor air quality standards"],

    ["pos-01", "Minimum Expectation", 0, "Grid connection", "Standard grid electricity connection"],
    ["pos-01", "Credit Achievement", 2, "Partial renewable", "30-60% renewable energy"],
    ["pos-01", "Exceptional Performance", 4, "Net zero energy", "100% renewable or net zero energy"],

    ["pos-02", "Minimum Expectation", 0, "Code compliance", "Meet minimum energy code requirements"],
    ["pos-02", "Credit Achievement", 2, "20% improvement", "20% better than code baseline"],
    ["pos-02", "Exceptional Performance", 4, "50% improvement", "50% better than code baseline"],

    ["pos-03", "Minimum Expectation", 0, "Carbon calculation", "Complete upfront carbon calculation"],
    ["pos-03", "Credit Achievement", 2, "20% reduction", "20% reduction from baseline"],
    ["pos-03", "Exceptional Performance", 4, "40% reduction", "40% reduction from baseline"],

    ["lead-01", "Minimum Expectation", 0, "Basic commitment", "Commit to Green Star certification"],
    ["lead-01", "Credit Achievement", 2, "Industry engagement", "Engage with industry on sustainability"],
    ["lead-01", "Exceptional Performance", 4, "Market leadership", "Demonstrate market transformation leadership"],
]

for row_idx, row_data in enumerate(levels_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_levels.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx in [4, 5] else Alignment(vertical='top')
        # Color code by level type
        if col_idx == 2:
            if value == "Minimum Expectation":
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue
            elif value == "Credit Achievement":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green
            elif value == "Exceptional Performance":
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber

ws_levels.column_dimensions['A'].width = 12
ws_levels.column_dimensions['B'].width = 25
ws_levels.column_dimensions['C'].width = 10
ws_levels.column_dimensions['D'].width = 30
ws_levels.column_dimensions['E'].width = 50
ws_levels.freeze_panes = 'A2'

# ============================================================
# SHEET 5: Criteria
# ============================================================
ws_criteria = wb.create_sheet("Criteria")
ws_criteria.sheet_properties.tabColor = "2D5A27"

criteria_headers = ["Credit ID", "Level Type", "Criteria #", "Criteria Text", "Is Mandatory?", "Evidence Type"]
for col, header in enumerate(criteria_headers, 1):
    cell = ws_criteria.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_dark
    cell.border = thin_border
    cell.alignment = center_alignment

# Add validations
ws_criteria.add_data_validation(level_validation)
level_validation.add('B2:B1000')

mandatory_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws_criteria.add_data_validation(mandatory_validation)
mandatory_validation.add('E2:E1000')

evidence_validation = DataValidation(type="list", formula1='"Document,Calculation,Drawing,Report,Certificate,Photo,Declaration,Other"', allow_blank=False)
ws_criteria.add_data_validation(evidence_validation)
evidence_validation.add('F2:F1000')

# Sample criteria data
criteria_data = [
    ["resp-01", "Minimum Expectation", 1, "Complete an Environmental Management Plan (EMP)", "Yes", "Document"],
    ["resp-01", "Minimum Expectation", 2, "Implement waste management procedures", "Yes", "Document"],
    ["resp-01", "Credit Achievement", 1, "Achieve 80% construction waste diversion from landfill", "Yes", "Report"],
    ["resp-01", "Credit Achievement", 2, "Use at least 3 sustainable material strategies", "No", "Document"],
    ["resp-01", "Exceptional Performance", 1, "Achieve 95% construction waste diversion", "Yes", "Report"],
    ["resp-01", "Exceptional Performance", 2, "Third-party verified environmental management", "Yes", "Certificate"],

    ["health-01", "Minimum Expectation", 1, "Meet minimum ventilation rates per AS 1668.2", "Yes", "Calculation"],
    ["health-01", "Minimum Expectation", 2, "Use low-VOC paints and adhesives", "Yes", "Declaration"],
    ["health-01", "Credit Achievement", 1, "Provide 50% more fresh air than code minimum", "Yes", "Calculation"],
    ["health-01", "Credit Achievement", 2, "Install CO2 monitoring in occupied spaces", "Yes", "Drawing"],
    ["health-01", "Exceptional Performance", 1, "Achieve WELL Air certification or equivalent", "Yes", "Certificate"],

    ["pos-01", "Minimum Expectation", 1, "Document current energy source mix", "Yes", "Document"],
    ["pos-01", "Credit Achievement", 1, "Source at least 30% of energy from renewables", "Yes", "Certificate"],
    ["pos-01", "Credit Achievement", 2, "Install on-site renewable energy generation", "No", "Document"],
    ["pos-01", "Exceptional Performance", 1, "Achieve 100% renewable energy or carbon neutral operations", "Yes", "Certificate"],
]

for row_idx, row_data in enumerate(criteria_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_criteria.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx == 4 else Alignment(vertical='top')

ws_criteria.column_dimensions['A'].width = 12
ws_criteria.column_dimensions['B'].width = 25
ws_criteria.column_dimensions['C'].width = 12
ws_criteria.column_dimensions['D'].width = 60
ws_criteria.column_dimensions['E'].width = 14
ws_criteria.column_dimensions['F'].width = 15
ws_criteria.freeze_panes = 'A2'

# ============================================================
# SHEET 6: Documentation
# ============================================================
ws_docs = wb.create_sheet("Documentation")
ws_docs.sheet_properties.tabColor = "7C3AED"

doc_headers = ["Credit ID", "Level Type", "Document Name", "Document Type", "Description", "Template Available?", "Template Link"]
for col, header in enumerate(doc_headers, 1):
    cell = ws_docs.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_purple
    cell.border = thin_border
    cell.alignment = center_alignment

# Add validations
level_val_docs = DataValidation(type="list", formula1='"Minimum Expectation,Credit Achievement,Exceptional Performance,All Levels"', allow_blank=False)
ws_docs.add_data_validation(level_val_docs)
level_val_docs.add('B2:B500')

doc_type_validation = DataValidation(type="list", formula1='"Report,Calculation,Drawing,Certificate,Declaration,Photo Evidence,Specification,Contract,Other"', allow_blank=False)
ws_docs.add_data_validation(doc_type_validation)
doc_type_validation.add('D2:D500')

template_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws_docs.add_data_validation(template_validation)
template_validation.add('F2:F500')

# Sample documentation data
docs_data = [
    ["resp-01", "All Levels", "Environmental Management Plan", "Report", "Comprehensive EMP covering construction phase", "Yes", "templates/EMP_template.docx"],
    ["resp-01", "All Levels", "Waste Management Report", "Report", "Monthly waste tracking and diversion rates", "Yes", "templates/waste_report.xlsx"],
    ["resp-01", "Credit Achievement", "Sustainable Materials Schedule", "Specification", "List of sustainable materials used", "No", ""],
    ["health-01", "All Levels", "Ventilation Calculations", "Calculation", "Fresh air rates per AS 1668.2", "Yes", "templates/ventilation_calc.xlsx"],
    ["health-01", "All Levels", "Low-VOC Product Declarations", "Declaration", "EPDs or manufacturer declarations", "No", ""],
    ["health-01", "Exceptional Performance", "WELL Certification", "Certificate", "WELL Air feature certification", "No", ""],
    ["pos-01", "Credit Achievement", "Renewable Energy Certificate", "Certificate", "GreenPower or LGC certificates", "No", ""],
    ["pos-01", "All Levels", "Energy Model Report", "Report", "Building energy simulation results", "Yes", "templates/energy_model.docx"],
]

for row_idx, row_data in enumerate(docs_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_docs.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx in [3, 5] else Alignment(vertical='top')

ws_docs.column_dimensions['A'].width = 12
ws_docs.column_dimensions['B'].width = 22
ws_docs.column_dimensions['C'].width = 35
ws_docs.column_dimensions['D'].width = 15
ws_docs.column_dimensions['E'].width = 45
ws_docs.column_dimensions['F'].width = 18
ws_docs.column_dimensions['G'].width = 30
ws_docs.freeze_panes = 'A2'

# ============================================================
# SHEET 7: Templates
# ============================================================
ws_templates = wb.create_sheet("Templates")
ws_templates.sheet_properties.tabColor = "D97706"

template_headers = ["Template ID", "Template Name", "Building Type", "Description", "Target Star Rating", "Recommended Credits (comma-separated)", "Is Active?"]
for col, header in enumerate(template_headers, 1):
    cell = ws_templates.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_amber
    cell.border = thin_border
    cell.alignment = center_alignment

# Add validations
building_type_validation = DataValidation(type="list", formula1='"Office,Residential,Retail,Education,Healthcare,Industrial,Mixed Use,Other"', allow_blank=False)
ws_templates.add_data_validation(building_type_validation)
building_type_validation.add('C2:C100')

star_validation = DataValidation(type="list", formula1='"4 Star,5 Star,6 Star"', allow_blank=False)
ws_templates.add_data_validation(star_validation)
star_validation.add('E2:E100')

active_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws_templates.add_data_validation(active_validation)
active_validation.add('G2:G100')

# Template data
templates_data = [
    ["tpl-office-5", "Commercial Office - 5 Star", "Office", "Standard commercial office targeting 5 Star rating with focus on energy and IEQ", "5 Star", "resp-01,resp-02,health-01,health-02,health-03,pos-01,pos-02,pos-03,lead-01", "Yes"],
    ["tpl-office-6", "Premium Office - 6 Star", "Office", "Premium office development targeting 6 Star with comprehensive sustainability", "6 Star", "resp-01,resp-02,resp-03,resp-04,health-01,health-02,health-03,health-04,health-05,pos-01,pos-02,pos-03,pos-04,pos-05,lead-01,lead-03", "Yes"],
    ["tpl-resi-5", "Residential - 5 Star", "Residential", "Multi-residential building targeting 5 Star", "5 Star", "resp-01,health-01,health-02,health-03,pos-01,pos-02,pos-05,people-01,people-03", "Yes"],
    ["tpl-retail-5", "Retail Centre - 5 Star", "Retail", "Shopping centre targeting 5 Star rating", "5 Star", "resp-01,health-01,pos-01,pos-02,pos-04,place-01,place-02", "Yes"],
    ["tpl-edu-5", "Education - 5 Star", "Education", "School or university building targeting 5 Star", "5 Star", "resp-01,health-01,health-02,health-03,health-04,pos-01,pos-02,people-01,people-03", "Yes"],
    ["tpl-health-6", "Healthcare - 6 Star", "Healthcare", "Hospital or healthcare facility targeting 6 Star", "6 Star", "resp-01,resp-02,resp-03,health-01,health-02,health-03,health-04,resil-01,pos-01,pos-02,pos-05,people-02,people-03", "Yes"],
]

for row_idx, row_data in enumerate(templates_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_templates.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx in [4, 6] else Alignment(vertical='top')

ws_templates.column_dimensions['A'].width = 15
ws_templates.column_dimensions['B'].width = 30
ws_templates.column_dimensions['C'].width = 15
ws_templates.column_dimensions['D'].width = 55
ws_templates.column_dimensions['E'].width = 18
ws_templates.column_dimensions['F'].width = 60
ws_templates.column_dimensions['G'].width = 12
ws_templates.freeze_panes = 'A2'

# ============================================================
# SHEET 8: Synergies
# ============================================================
ws_synergies = wb.create_sheet("Synergies")
ws_synergies.sheet_properties.tabColor = "0D9488"

synergy_headers = ["Synergy ID", "Credit 1", "Credit 2", "Synergy Type", "Description", "Recommendation"]
for col, header in enumerate(synergy_headers, 1):
    cell = ws_synergies.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill_teal
    cell.border = thin_border
    cell.alignment = center_alignment

synergy_type_validation = DataValidation(type="list", formula1='"Strong,Moderate,Documentation"', allow_blank=False)
ws_synergies.add_data_validation(synergy_type_validation)
synergy_type_validation.add('D2:D200')

synergies_data = [
    ["syn-001", "pos-01", "pos-02", "Strong", "Energy source and energy use are directly related", "Pursuing both credits together maximizes energy strategy effectiveness"],
    ["syn-002", "pos-02", "pos-03", "Strong", "Energy efficiency reduces both operational and embodied carbon", "Consider holistic carbon strategy addressing both credits"],
    ["syn-003", "health-01", "health-02", "Moderate", "Ventilation strategy affects both air quality and thermal comfort", "Coordinate HVAC design to address both credits"],
    ["syn-004", "health-02", "health-03", "Moderate", "Lighting and thermal comfort both affect occupant satisfaction", "Consider integrated comfort strategy"],
    ["syn-005", "resp-01", "nature-01", "Moderate", "Construction practices directly impact site ecology", "Align EMP with ecology protection measures"],
    ["syn-006", "resp-02", "pos-02", "Documentation", "Metering supports energy use verification", "Same energy model can support both credits"],
    ["syn-007", "place-01", "people-01", "Moderate", "Transport and social procurement both support community", "Consider integrated community benefit strategy"],
    ["syn-008", "resil-01", "resil-02", "Strong", "Climate resilience includes flood risk management", "Develop comprehensive climate risk assessment"],
]

for row_idx, row_data in enumerate(synergies_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_synergies.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment if col_idx in [5, 6] else Alignment(vertical='top')
        if col_idx == 4:
            if value == "Strong":
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif value == "Moderate":
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif value == "Documentation":
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

ws_synergies.column_dimensions['A'].width = 12
ws_synergies.column_dimensions['B'].width = 12
ws_synergies.column_dimensions['C'].width = 12
ws_synergies.column_dimensions['D'].width = 15
ws_synergies.column_dimensions['E'].width = 50
ws_synergies.column_dimensions['F'].width = 55
ws_synergies.freeze_panes = 'A2'

# ============================================================
# Save the workbook
# ============================================================
wb.save('/home/user/Submissiong/GreenStar_Admin_Data.xlsx')
print("✓ Admin spreadsheet created: GreenStar_Admin_Data.xlsx")
print("\nSheets created:")
print("  1. Instructions - How to use the spreadsheet")
print("  2. Categories - 8 category definitions")
print("  3. Credits - All credits with dropdowns")
print("  4. Levels - Performance levels per credit")
print("  5. Criteria - Detailed criteria with evidence types")
print("  6. Documentation - Required documents")
print("  7. Templates - Project templates")
print("  8. Synergies - Credit relationships")
